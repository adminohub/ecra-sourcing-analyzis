import streamlit as st
import pandas as pd
import plotly.graph_objects as go_fig
from datetime import datetime
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

st.set_page_config(page_title="ECRA — Analyse Produit", page_icon="🐾",
                   layout="wide", initial_sidebar_state="expanded")

NAVY  = "#1E2761"; GOLD  = "#D4A843"; GREEN = "#1A7A3A"
DGRAY = "#2C3E50"; RED   = "#C0392B"; LIGHT = "#F0F2F8"

SOUS_NICHES = ["Accessoires de Promenade","Jouets Intelligents","Confort & Bien-être",
                "Toilettage & Hygiène","Éducation & Comportement","Vêtements & Style",
                "Alimentation & Distribution"]

SCORING_CRITERIA = [
    ("demande_marche",         "Demande marché",         20, 4),
    ("potentiel_pub",          "Potentiel pub",           15, 3),
    ("marge_brute",            "Marge brute",             15, 3),
    ("momentum_tendance",      "Momentum tendance",       10, 2),
    ("saturation",             "Saturation marché",       10, 2),
    ("faisabilite_logistique", "Faisabilité logistique",  10, 2),
    ("differenciation",        "Différenciation",          5, 1),
    ("brandabilite",           "Brandabilité",             5, 1),
    ("scalabilite",            "Scalabilité",              5, 1),
    ("private_label",          "Potentiel Private Label",  5, 1),
]

RUBRIC = {
    "demande_marche":         [(4,"Niche de niche, besoin anecdotique"),(8,"Besoin réel mais volume faible"),(12,"Demande régionale ou saisonnière"),(16,"Forte demande nationale stable"),(20,"Evergreen, volume élevé, achat répété")],
    "potentiel_pub":          [(3,"Difficile à filmer, angle faible"),(6,"Un seul angle, peu accrocheur"),(9,"1-2 angles corrects, démo possible"),(12,"2-3 hooks forts, vidéo UGC facile"),(15,"Avant/après saisissant, viral potentiel")],
    "marge_brute":            [(3,"< 2x le coût (zone de risque)"),(6,"2x le coût (marge minimale)"),(9,"3x le coût (standard dropshipping)"),(12,"4x le coût (bonne marge)"),(15,"5x+ le coût (excellente marge)")],
    "momentum_tendance":      [(2,"Tendance plate ou déclinante"),(4,"Légère hausse récente"),(6,"Montée stable sur 6 mois"),(8,"Accélération nette sur 12 mois"),(10,"Courbe exponentielle, buzz actuel")],
    "saturation":             [(2,"Marché ultra-saturé, < 5 angles libres"),(4,"Marché chargé, quelques niches"),(6,"Concurrence normale, espaces libres"),(8,"Peu de concurrents sérieux"),(10,"Marché quasi-vierge, first-mover")],
    "faisabilite_logistique": [(2,"Délai > 30j, fournisseur instable"),(4,"Délai 15-30j, fournisseur moyen"),(6,"Délai 10-15j, CJ/AutoDS dispo"),(8,"Délai < 10j, stock stable"),(10,"EU < 7j, stock garanti")],
    "differenciation":        [(1,"Produit générique identique"),(2,"Légère variation couleur/taille"),(3,"Bundle ou packaging différenciant"),(4,"Angle unique, storytelling propre"),(5,"USP forte, innovation visible")],
    "brandabilite":           [(1,"Nom ingrat, pas d'univers"),(2,"Produit fonctionnel, peu émotionnel"),(3,"Univers possible avec effort créatif"),(4,"Nom court, visuel fort"),(5,"Marque évidente, fidélisation naturelle")],
    "scalabilite":            [(1,"Produit unique, pas d'extension"),(2,"1-2 complémentaires max"),(3,"Gamme 3-5 produits envisageable"),(4,"Sous-niche complète adressable"),(5,"Scalable, subscription possible")],
    "private_label":          [(1,"Produit générique ou breveté"),(2,"Possible mais peu rentable"),(3,"Envisageable à 30+ cmd/mois"),(4,"Fournisseur OEM identifié"),(5,"Label propre dès 50 cmd/mois")],
}

VERDICT_COLORS = {
    "🚀 TEST PRIORITAIRE":"1A7A3A","✅ TEST PRUDENT":"2980B9",
    "⏳ SURVEILLER":"E67E22","❌ NE PAS TESTER":"C0392B",
    "🔴 ABANDONNER":"7F8C8D","🚫 BUDGET PUB BLOQUÉ":"C0392B",
}

def banner(text, bg=NAVY):
    st.markdown(
        f"<div style='background:{bg};color:white;padding:10px 16px;"
        f"border-radius:8px;font-weight:700;margin-bottom:1rem;font-size:1rem;'>"
        f"{text}</div>", unsafe_allow_html=True)

def verdict_info(total, has_nogo):
    if has_nogo: return "🚫 BUDGET PUB BLOQUÉ", RED, "Un ou plusieurs outils ont rendu un NO-GO."
    if total >= 80: return "🚀 TEST PRIORITAIRE", GREEN, "Lancer immédiatement. Mobiliser budget TikTok Ads."
    if total >= 65: return "✅ TEST PRUDENT", "#2980B9", "Budget réduit. Surveiller ROAS à J+7."
    if total >= 50: return "⏳ SURVEILLER", "#E67E22", "Ne pas lancer. Ré-évaluer dans 30 jours."
    if total >= 35: return "❌ NE PAS TESTER", RED, "Retourner à la recherche produit."
    return "🔴 ABANDONNER", "#7F8C8D", "Exclure définitivement. Informer le vendeur."

# Session defaults
PROD_DEF = dict(
    produit="", lien_fournisseur="", sous_niche=SOUS_NICHES[0],
    probleme="", cible="", benefice="",
    gt_go=None, gt_kw1="", gt_kw2="", gt_note="",
    bsr_go=None, bsr_note="",
    wh_go=None,  wh_note="",
    minea_go=None, minea_note="",
    scores={k: 0 for k, *_ in SCORING_CRITERIA},
    source="", commentaire="", trends_data=None,
)
if "etape"          not in st.session_state: st.session_state.etape = 1
if "vendeur"        not in st.session_state: st.session_state.vendeur = ""
if "produits_liste" not in st.session_state: st.session_state.produits_liste = []
for k, v in PROD_DEF.items():
    if k not in st.session_state:
        st.session_state[k] = dict(v) if isinstance(v, dict) else v

def reset_product():
    for k, v in PROD_DEF.items():
        st.session_state[k] = dict(v) if isinstance(v, dict) else v
    st.session_state.etape = 1

def nav(n): st.session_state.etape = n; st.rerun()

# CSS
st.markdown(f"""<style>
/* Cache la barre du haut (Deploy, menu) */
header[data-testid="stHeader"] {{display: none !important;}}
#MainMenu {{display: none !important;}}
footer {{display: none !important;}}
/* Descend le contenu pour ne pas coller au bord */
.block-container {{
    padding-top: 2rem !important;
    padding-bottom: 2rem !important;
}}
.main{{background:#F8F9FC}}
h1,h2,h3{{color:{NAVY}}}
.stButton>button{{background:{NAVY};color:white;border-radius:6px;border:none;padding:.45rem 1.4rem;font-weight:600}}
.stButton>button:hover{{background:{GOLD};color:{NAVY}}}
</style>""", unsafe_allow_html=True)

# ── SIDEBAR ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"<div style='background:{NAVY};padding:14px;border-radius:8px;text-align:center;margin-bottom:8px;'><span style='color:{GOLD};font-size:1.3rem;font-weight:800;'>🐾 ECRA</span><br><span style='color:#AAB4D4;font-size:.75rem;'>Analyse Produit Sourcing</span></div>", unsafe_allow_html=True)
    st.session_state.vendeur = st.text_input("👤 Nom du vendeur", st.session_state.vendeur, placeholder="ex: Jean Dupont")
    st.markdown("---")
    st.markdown(f"<p style='color:{GOLD};font-weight:700;margin-bottom:4px;'>ÉTAPES</p>", unsafe_allow_html=True)
    for num, label in [(1,"① Identification"),(2,"② Google Trends"),(3,"③ Amazon BSR"),(4,"④ WinningHunter"),(5,"⑤ Minea"),(6,"⑥ Scoring /100"),(7,"⑦ Résultat & Export")]:
        done = st.session_state.etape > num; active = st.session_state.etape == num
        color = GOLD if active else (GREEN if done else "#999")
        icon  = "✅" if done else ("▶" if active else "○")
        st.markdown(f"<p style='color:{color};margin:2px 0;font-size:.88rem;'>{icon} {label}</p>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown(f"<p style='color:{GOLD};font-weight:700;margin-bottom:4px;'>PRÉ-SCREEN</p>", unsafe_allow_html=True)
    for name, key in [("Google Trends","gt_go"),("Amazon BSR","bsr_go"),("WinningHunter","wh_go"),("Minea","minea_go")]:
        val = st.session_state[key]
        color = GREEN if val is True else (RED if val is False else "#888")
        label = "GO ✅" if val is True else ("NO-GO ❌" if val is False else "—")
        st.markdown(f"<span style='background:{color};color:white;padding:2px 8px;border-radius:12px;font-size:.8rem;font-weight:700;'>{label}</span> {name}", unsafe_allow_html=True)
    st.markdown("---")
    total_sb = sum(st.session_state.scores.values())
    st.markdown(f"<p style='color:{GOLD};font-weight:700;margin-bottom:2px;'>SCORE ACTUEL</p><p style='font-size:1.6rem;font-weight:800;color:{NAVY};margin:0;'>{total_sb}<span style='font-size:1rem;color:#888;'>/100</span></p>", unsafe_allow_html=True)
    st.progress(total_sb / 100)
    st.markdown("---")
    nb = len(st.session_state.produits_liste)
    st.markdown(f"<p style='color:{GOLD};font-weight:700;margin-bottom:4px;'>SESSION ({nb} produit{'s' if nb>1 else ''})</p>", unsafe_allow_html=True)
    if nb == 0:
        st.markdown("<p style='color:#999;font-size:.82rem;'>Aucun produit encore ajouté.</p>", unsafe_allow_html=True)
    for i, p in enumerate(st.session_state.produits_liste, 1):
        col = GREEN if any(x in p["verdict"] for x in ["PRIORITAIRE","PRUDENT"]) else RED
        st.markdown(f"<p style='font-size:.82rem;margin:2px 0;'><b style='color:{NAVY};'>#{i}</b> {p['produit'][:20]}<br><span style='color:{col};font-weight:700;'>{p['score']}/100</span></p>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
# ÉTAPE 1
# ═══════════════════════════════════════════════════════════════════
if st.session_state.etape == 1:
    banner("A — IDENTIFICATION DU PRODUIT")
    st.caption("Remplis les informations de base soumises par le vendeur.")
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.produit = st.text_input("Nom du produit *", st.session_state.produit, placeholder="ex: Nettoyeur de pattes pour chien")
        st.session_state.sous_niche = st.selectbox("Sous-niche *", SOUS_NICHES, index=SOUS_NICHES.index(st.session_state.sous_niche))
        st.session_state.probleme = st.text_area("Problème résolu *", st.session_state.probleme, height=80, placeholder="ex: Pattes sales après sortie")
    with col2:
        st.session_state.lien_fournisseur = st.text_input("Lien fournisseur", st.session_state.lien_fournisseur, placeholder="https://...")
        st.session_state.cible = st.text_input("Cible client", st.session_state.cible, placeholder="ex: Maîtres de chiens, 25-45 ans")
        st.session_state.benefice = st.text_area("Bénéfice principal", st.session_state.benefice, height=80, placeholder="ex: Évite de salir la maison")
    st.markdown("---")
    if st.button("Suivant → Google Trends", type="primary"):
        if not st.session_state.produit: st.error("Le nom du produit est obligatoire.")
        else: nav(2)

# ═══════════════════════════════════════════════════════════════════
# ÉTAPE 2
# ═══════════════════════════════════════════════════════════════════
elif st.session_state.etape == 2:
    banner("B — ① GOOGLE TRENDS — Tendance favorable ?")
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.gt_kw1 = st.text_input("Mot-clé Trends #1 — français *", st.session_state.gt_kw1, placeholder="ex: collier anti aboiement")
    with col2:
        st.session_state.gt_kw2 = st.text_input("Mot-clé Trends #2 — anglais *", st.session_state.gt_kw2, placeholder="ex: anti bark collar")
    if st.button("🔍 Récupérer les données Google Trends"):
        if not st.session_state.gt_kw1 or not st.session_state.gt_kw2:
            st.error("Les deux mots-clés sont obligatoires — un en français, un en anglais.")
        else:
            with st.spinner("Connexion à Google Trends..."):
                try:
                    from pytrends.request import TrendReq
                    pt = TrendReq(hl="fr-FR", tz=360, timeout=(10,25), retries=2, backoff_factor=0.5)
                    kws = [st.session_state.gt_kw1.strip(), st.session_state.gt_kw2.strip()]
                    pt.build_payload(kws, cat=0, timeframe="today 12-m", geo="FR", gprop="")
                    df = pt.interest_over_time().drop(columns=["isPartial"], errors="ignore")
                    if df.empty: st.warning("Aucune donnée. Essaie des termes plus génériques."); st.session_state.trends_data = None
                    else: st.session_state.trends_data = df; st.success("Données récupérées !")
                except Exception as err:
                    st.warning(f"Google Trends indisponible : {str(err)[:120]}\n\nUtilise le lien manuel ci-dessous.")
                    st.session_state.trends_data = None
    if st.session_state.trends_data is not None:
        df = st.session_state.trends_data
        fig = go_fig.Figure()
        for i, col in enumerate(df.columns):
            fig.add_trace(go_fig.Scatter(x=df.index, y=df[col], name=col,
                line=dict(color=[NAVY,GOLD][i%2], width=2.5),
                fill="tozeroy" if i==0 else None, fillcolor="rgba(30,39,97,0.08)"))
        h1 = df.iloc[:len(df)//2].mean().mean(); h2 = df.iloc[len(df)//2:].mean().mean()
        fig.update_layout(title="Intérêt (France, 12 mois)", xaxis_title="Date",
            yaxis_title="Intérêt (0-100)", plot_bgcolor="white", paper_bgcolor="white", height=320,
            legend=dict(orientation="h", yanchor="bottom", y=1.02))
        st.plotly_chart(fig, use_container_width=True)
        if h2 > h1 * 1.05: st.success(f"📈 Tendance montante — 2e semestre +{((h2/h1)-1)*100:.0f}% vs 1er")
        else: st.warning("📉 Tendance plate ou déclinante")
    with st.expander("📖 Vérification manuelle (si fetch échoue)"):
        kw1 = st.session_state.gt_kw1.strip() if st.session_state.gt_kw1 else ""
        kw2 = st.session_state.gt_kw2.strip() if st.session_state.gt_kw2 else ""
        if kw1:
            q = (kw1 + ("," + kw2 if kw2 else "")).replace(" ", "+")
            url = f"https://trends.google.fr/trends/explore?q={q}&geo=FR&date=today+12-m"
            st.markdown(f"[🔗 Ouvrir Google Trends]({url})")
            st.caption("Si tu vois 'Pas assez de données' : utilise un terme plus court et plus générique (ex: 'collier chien' au lieu de 'collier anti aboiement chien électrique').")
        st.markdown("""
**GO ✅ si :** courbe montante ou stable sur 12 mois.  
**NO-GO ❌ si :** courbe en baisse ou donnée insuffisante sur le terme générique.
        """)
    st.session_state.gt_note = st.text_area("📝 Note", st.session_state.gt_note, height=55, placeholder="ex: Hausse depuis oct 2025...")
    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ GO TENDANCE", use_container_width=True): st.session_state.gt_go = True; nav(3)
    with c2:
        if st.button("❌ NO-GO (tendance défavorable)", use_container_width=True): st.session_state.gt_go = False; nav(3)
    st.button("← Retour", on_click=lambda: nav(1))

# ═══════════════════════════════════════════════════════════════════
# ÉTAPE 3
# ═══════════════════════════════════════════════════════════════════
elif st.session_state.etape == 3:
    banner("B — ② AMAZON BEST SELLERS — Demande prouvée ?")
    pq = (st.session_state.produit or "produit").replace(" ", "+")
    c1, c2 = st.columns(2)
    with c1: st.markdown(f"[🔗 Amazon.fr](https://www.amazon.fr/s?k={pq}&s=review-rank)")
    with c2: st.markdown(f"[🔗 Amazon.de](https://www.amazon.de/s?k={pq}&s=review-rank)")
    st.markdown("---")
    c1 = st.checkbox("Produit dans le Top 100 BSR de sa catégorie")
    c2 = st.checkbox("Plusieurs vendeurs ont + de 50 avis")
    c3 = st.checkbox("Vendu en France, Belgique ou Allemagne")
    nb_ok = sum([c1,c2,c3])
    if nb_ok==3: st.success("✅ Tous les signaux verts")
    elif nb_ok>=1: st.warning(f"⚠️ {nb_ok}/3 signaux positifs")
    else: st.error("❌ Demande non prouvée")
    st.session_state.bsr_note = st.text_area("📝 Note BSR", st.session_state.bsr_note, height=55, placeholder="ex: Top 50 BSR, ~200 avis/vendeur")
    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ GO DEMANDE", use_container_width=True): st.session_state.bsr_go = True; nav(4)
    with c2:
        if st.button("❌ NO-GO (demande non prouvée)", use_container_width=True): st.session_state.bsr_go = False; nav(4)
    st.button("← Retour", on_click=lambda: nav(2))

# ═══════════════════════════════════════════════════════════════════
# ÉTAPE 4
# ═══════════════════════════════════════════════════════════════════
elif st.session_state.etape == 4:
    banner("B — ③ WINNINGHUNTER — Concurrence & angles ?")
    st.markdown("[🔗 Ouvrir WinningHunter](https://winninghunter.com)")

    with st.expander("📋 Filtres recommandés à appliquer dans WinningHunter", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"""
<div style='background:#EEF0FA;border-left:4px solid {GOLD};padding:10px 14px;border-radius:4px;color:#1E2761;'>
<b style='color:#1E2761;'>🌍 Pays cibles</b><br>
<span style='color:#2C3E50;'>France · Belgique · Suisse</span><br><br>
<b style='color:#1E2761;'>📅 Période</b><br>
<span style='color:#2C3E50;'>30 derniers jours (vue récente)</span><br><br>
<b style='color:#1E2761;'>🏷️ Catégorie</b><br>
<span style='color:#2C3E50;'>Pets & Animals → Dogs</span>
</div>""", unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
<div style='background:#EEF0FA;border-left:4px solid {NAVY};padding:10px 14px;border-radius:4px;color:#1E2761;'>
<b style='color:#1E2761;'>💰 Fourchette de prix</b><br>
<span style='color:#2C3E50;'>15€ – 80€ (dropshipping EU rentable)</span><br><br>
<b style='color:#1E2761;'>📦 Type de store</b><br>
<span style='color:#2C3E50;'>Shopify uniquement</span><br><br>
<b style='color:#1E2761;'>🔢 Nb de boutiques min.</b><br>
<span style='color:#2C3E50;'>Laisser à 1 (voir toute la concurrence)</span>
</div>""", unsafe_allow_html=True)
        st.markdown(f"""
<div style='background:#FFF9E6;border-left:4px solid {GOLD};padding:10px 14px;border-radius:4px;margin-top:10px;'>
<b style='color:#1E2761;'>💡 Comment lire le ratio prix ?</b><br>
<span style='color:#2C3E50;'><b>Ton sourcing</b> = Prix fournisseur CJ/AutoDS + livraison EU (déjà sur la fiche produit).<br>
<b>Prix concurrent</b> = Prix de vente affiché sur les boutiques Shopify dans WinningHunter.<br><br>
<b>Exemple :</b> CJ te coûte <b>8€</b> tout inclus → concurrent vend à <b>34€</b> → ratio = 34÷8 = <b>4,25x ✅</b></span><br>
<span style='color:{RED};font-weight:600;'>Si le ratio est inférieur à 2x, la marge est insuffisante même avant les frais pub.</span>
</div>""", unsafe_allow_html=True)

    st.markdown("---")
    q1 = st.radio("Saturation de la niche ?", ["🔴 Saturée — plus de 20 boutiques actives","🟡 Chargée — 10-20 boutiques","🟢 Normale — moins de 10 boutiques","🟢 Vierge — moins de 5 boutiques sérieuses"], key="wh_q1")
    q2 = st.radio("Angles marketing disponibles ?", ["❌ Tous les angles exploités massivement","⚠️ 1-2 angles libres mais faibles","✅ 2-3 angles corrects identifiés","✅ 3+ angles forts, dont un jamais vu"], key="wh_q2")

    st.markdown(f"""
<div style='background:#F0F2F8;padding:8px 12px;border-radius:6px;font-size:.85rem;margin-bottom:8px;'>
<b>📌 Rappel ratio :</b> Coût total fournisseur (CJ/AutoDS livraison incluse) → cherche ce prix dans WinningHunter sur les boutiques concurrentes → divise prix vente ÷ coût.
</div>""", unsafe_allow_html=True)
    q3 = st.radio("Prix concurrents vs ton sourcing ?", ["❌ Concurrents < 2x ton coût","⚠️ Concurrents à 2-3x ton coût","✅ Concurrents à 3-5x ton coût","✅ Concurrents à 5x+ ton coût"], key="wh_q3")
    good = sum(["🔴" not in q1, "❌" not in q2, "❌" not in q3])
    if good==3: st.success("✅ Tous les signaux verts")
    elif good==2: st.warning("⚠️ Signaux mitigés")
    else: st.error("❌ Concurrence trop difficile")
    st.session_state.wh_note = st.text_area("📝 Angles / observations", st.session_state.wh_note, height=55)
    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ GO — Concurrence exploitable", use_container_width=True): st.session_state.wh_go = True; nav(5)
    with c2:
        if st.button("❌ NO-GO (saturé / aucun angle)", use_container_width=True): st.session_state.wh_go = False; nav(5)
    st.button("← Retour", on_click=lambda: nav(3))

# ═══════════════════════════════════════════════════════════════════
# ÉTAPE 5
# ═══════════════════════════════════════════════════════════════════
elif st.session_state.etape == 5:
    banner("B — ④ MINEA — Angle publicitaire vierge ?")
    st.markdown("[🔗 Ouvrir Minea](https://minea.com)")

    with st.expander("📋 Filtres recommandés à appliquer dans Minea", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"""
<div style='background:#EEF0FA;border-left:4px solid {GOLD};padding:10px 14px;border-radius:4px;'>
<b style='color:#1E2761;'>📱 Réseaux à analyser</b><br>
<span style='color:#2C3E50;'>TikTok Ads <i>(priorité absolue)</i><br>
Facebook/Meta Ads <i>(complémentaire)</i></span><br><br>
<b style='color:#1E2761;'>🌍 Pays cibles</b><br>
<span style='color:#2C3E50;'>France · Belgique · Suisse</span><br><br>
<b style='color:#1E2761;'>📅 Période</b><br>
<span style='color:#2C3E50;'>30 derniers jours pour les pubs actives</span>
</div>""", unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
<div style='background:#EEF0FA;border-left:4px solid {NAVY};padding:10px 14px;border-radius:4px;'>
<b style='color:#1E2761;'>🔍 Recherche</b><br>
<span style='color:#2C3E50;'>Tape le nom en français ET en anglais</span><br><br>
<b style='color:#1E2761;'>📊 Trier par</b><br>
<span style='color:#2C3E50;'>"Récent" d'abord → puis "Engagements"</span><br><br>
<b style='color:#1E2761;'>🏷️ Filtre statut pub</b><br>
<span style='color:#2C3E50;'>"En cours" = pubs actives avec budget en vie</span>
</div>""", unsafe_allow_html=True)
        st.markdown(f"""
<div style='background:#FFF9E6;border-left:4px solid {GOLD};padding:10px 14px;border-radius:4px;margin-top:10px;'>
<b style='color:#1E2761;'>💡 Ce que tu cherches concrètement :</b><br>
<span style='color:#2C3E50;'><b>Angle vierge</b> = un hook créatif que personne n'utilise encore massivement.<br><br>
<b>Exemples :</b> "chien senior qui souffre des articulations" · "maîtresse débordée après promenade sous la pluie" · "routine bain maison filmée en UGC"</span><br>
<span style='color:{RED};font-weight:600;'>Si toutes les pubs se ressemblent → marché pub saturé = NO-GO.</span>
</div>""", unsafe_allow_html=True)

    st.markdown("---")
    q1 = st.radio("Volume de pubs actives ?", ["🔴 Massif — centaines de pubs","🟡 Modéré — 20-100 pubs","🟢 Faible — moins de 20 pubs","🟢 Quasi-nul — moins de 5 pubs"], key="minea_q1")
    q2 = st.radio("Angles créatifs déjà utilisés ?", ["❌ Tous les angles évidents saturés","⚠️ Angles principaux pris, variantes possibles","✅ 1-2 angles forts encore libres","✅ Angle vraiment vierge, jamais vu en pub"], key="minea_q2")
    q3 = st.radio("Performance des pubs existantes ?", ["❌ Pubs depuis > 6 mois avec gros budgets","⚠️ Pubs récentes mais concurrents actifs","✅ Pubs récentes, peu d'engagement","✅ Aucune pub établie"], key="minea_q3")
    good = sum(["🔴" not in q1, "❌" not in q2, "❌" not in q3])
    if good==3: st.success("✅ Angle vierge confirmé")
    elif good==2: st.warning("⚠️ Angle partiellement libre")
    else: st.error("❌ Espace pub trop occupé")
    st.session_state.minea_note = st.text_area("📝 Angle vierge / observations", st.session_state.minea_note, height=55)
    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ GO — Angle vierge identifié", use_container_width=True): st.session_state.minea_go = True; nav(6)
    with c2:
        if st.button("❌ NO-GO (espace pub saturé)", use_container_width=True): st.session_state.minea_go = False; nav(6)
    st.button("← Retour", on_click=lambda: nav(4))

# ═══════════════════════════════════════════════════════════════════
# ÉTAPE 6
# ═══════════════════════════════════════════════════════════════════
elif st.session_state.etape == 6:
    banner("C — SCORING PONDÉRÉ /100")
    gos = [st.session_state.gt_go, st.session_state.bsr_go, st.session_state.wh_go, st.session_state.minea_go]
    if any(g is False for g in gos): st.error("⚠️ Budget pub BLOQUÉ — scoring informatif uniquement.")
    else: st.success("✅ Pré-screen validé.")
    st.caption("Sélectionne le niveau pour chaque critère.")
    st.markdown("---")
    total = 0
    for key, label, mx, step in SCORING_CRITERIA:
        rubric  = RUBRIC[key]
        options = [0] + [v for v, _ in rubric]
        labels  = ["— Non évalué"] + [f"{v} pts — {desc}" for v, desc in rubric]
        st.markdown(f"**{label}** */{mx} pts*")
        cur = st.session_state.scores.get(key, 0)
        idx = options.index(cur) if cur in options else 0
        choice = st.selectbox(f"_{label}_", labels, index=idx, key=f"sel_{key}", label_visibility="collapsed")
        chosen = options[labels.index(choice)]
        st.session_state.scores[key] = chosen; total += chosen
        pct = chosen / mx if mx else 0
        bar_c = GREEN if pct>=.8 else ("#E67E22" if pct>=.5 else RED)
        st.markdown(f"<div style='background:#E0E4F0;border-radius:4px;height:5px;margin-bottom:12px;'><div style='width:{pct*100:.0f}%;background:{bar_c};height:5px;border-radius:4px;'></div></div>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown(f"### Score total : **{total}/100**")
    st.progress(total / 100)
    c1, c2 = st.columns(2)
    with c1: st.button("← Retour", on_click=lambda: nav(5))
    with c2:
        if st.button("Suivant → Résultat & Export", type="primary"): nav(7)

# ═══════════════════════════════════════════════════════════════════
# ÉTAPE 7
# ═══════════════════════════════════════════════════════════════════
elif st.session_state.etape == 7:
    banner("D — RÉSULTAT FINAL & EXPORT EXCEL")
    total    = sum(st.session_state.scores.values())
    gos      = [st.session_state.gt_go, st.session_state.bsr_go, st.session_state.wh_go, st.session_state.minea_go]
    has_nogo = any(g is False for g in gos)
    verdict, v_color, v_msg = verdict_info(total, has_nogo)

    # Score card
    _, mid, _ = st.columns([1,2,1])
    with mid:
        st.markdown(f"<div style='background:{NAVY};color:white;border-radius:12px;padding:1.5rem;text-align:center;'><div style='font-size:3.5rem;font-weight:800;color:{GOLD};'>{total}</div><div style='color:#AAB4D4;'>points sur 100</div><div style='font-size:1.3rem;font-weight:700;margin-top:8px;'>{verdict}</div><div style='color:#CCD0E0;font-size:.85rem;margin-top:4px;'>{v_msg}</div></div>", unsafe_allow_html=True)

    st.markdown("---")

    # Radar chart — using go_fig (alias for plotly.graph_objects, never shadowed)
    cat_labels = [label for _, label, _, _ in SCORING_CRITERIA]
    values     = [st.session_state.scores[k] for k, *_ in SCORING_CRITERIA]
    maxes      = [mx for _, _, mx, _ in SCORING_CRITERIA]
    pcts       = [v/m*100 for v,m in zip(values, maxes)]
    radar = go_fig.Figure(go_fig.Scatterpolar(
        r=pcts+[pcts[0]], theta=cat_labels+[cat_labels[0]],
        fill="toself", fillcolor="rgba(30,39,97,0.15)",
        line=dict(color=NAVY, width=2)))
    radar.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0,100])),
        showlegend=False, height=400, paper_bgcolor="white",
        title=dict(text="Profil du produit (% du max)", font=dict(color=NAVY)))
    st.plotly_chart(radar, use_container_width=True)

    rows = [{"Critère":label,"Score":f"{st.session_state.scores[k]}/{mx}","%":f"{st.session_state.scores[k]/mx*100:.0f}%"}
            for k,label,mx,_ in SCORING_CRITERIA]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1: st.session_state.source = st.text_input("Source de validation", st.session_state.source, placeholder="Google Trends ✓ | Amazon BSR ✓...")
    with c2: st.session_state.commentaire = st.text_area("Commentaire", st.session_state.commentaire, height=70)

    # Build current product dict
    go_str = {True:"GO ✅", False:"NO-GO ❌", None:"—"}
    pre_sc = "✅ PASS" if all(g is True for g in gos) else ("🚫 Budget pub bloqué" if has_nogo else "⏳ En cours…")
    cur_p  = dict(
        produit=st.session_state.produit, lien_fournisseur=st.session_state.lien_fournisseur,
        sous_niche=st.session_state.sous_niche, probleme=st.session_state.probleme,
        cible=st.session_state.cible, benefice=st.session_state.benefice,
        gt=go_str[st.session_state.gt_go], gt_kw1=st.session_state.gt_kw1, gt_kw2=st.session_state.gt_kw2,
        bsr=go_str[st.session_state.bsr_go], wh=go_str[st.session_state.wh_go],
        minea=go_str[st.session_state.minea_go], pre_screen=pre_sc,
        **{k: st.session_state.scores[k] for k,*_ in SCORING_CRITERIA},
        score=total, verdict=verdict, source=st.session_state.source,
        commentaire=st.session_state.commentaire, date=datetime.now().strftime("%d/%m/%Y %H:%M"),
    )

    st.markdown("---")
    banner("GESTION SESSION VENDEUR", DGRAY)

    col_add, col_exp = st.columns(2)

    with col_add:
        if st.button("➕ Ajouter ce produit & analyser le suivant", type="primary", use_container_width=True):
            existing = [p["produit"] for p in st.session_state.produits_liste]
            if st.session_state.produit in existing:
                st.session_state.produits_liste[existing.index(st.session_state.produit)] = cur_p
                st.success(f"✅ **{st.session_state.produit}** mis à jour.")
            else:
                st.session_state.produits_liste.append(cur_p)
                st.success(f"✅ **{st.session_state.produit}** ajouté — {len(st.session_state.produits_liste)} produit(s) en session.")
            reset_product()

    with col_exp:
        all_prods = list(st.session_state.produits_liste)
        already   = [p["produit"] for p in all_prods]
        if st.session_state.produit and st.session_state.produit not in already:
            all_prods.append(cur_p)

        if st.button(f"📥 Exporter Excel — {len(all_prods)} produit(s)", use_container_width=True, disabled=len(all_prods)==0):
            def fx(h): return PatternFill("solid", start_color=h, fgColor=h)
            def fb(bold=False,color="FFFFFF",size=9): return Font(name="Arial",bold=bold,color=color,size=size)
            def fd():
                s=Side(style="thin",color="BBBBBB"); return Border(left=s,right=s,top=s,bottom=s)
            def fc(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
            def fl(): return Alignment(horizontal="left",  vertical="center",wrap_text=True)

            wb2 = Workbook(); ws2 = wb2.active; ws2.title = "Scoring"
            ws2.sheet_view.showGridLines = False
            vnd = st.session_state.vendeur or "Vendeur"

            ws2.merge_cells("A1:AB1")
            ws2["A1"].value = f"RAPPORT D'ANALYSE — {vnd} | Niche Chien 2026 | ECRA"
            ws2["A1"].font = fb(True,size=13); ws2["A1"].fill = fx("1E2761"); ws2["A1"].alignment = fc()
            ws2.row_dimensions[1].height = 34

            ws2.merge_cells("A2:AB2")
            ws2["A2"].value = f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')} | {len(all_prods)} produit(s)"
            ws2["A2"].font = Font(name="Arial",italic=True,color="D4A843",size=9)
            ws2["A2"].fill = fx("1E2761"); ws2["A2"].alignment = fl()
            ws2.row_dimensions[2].height = 18

            for rng,lbl,bg,fg in [("A3:F3","A — IDENTIFICATION","1E2761","FFFFFF"),
                                    ("G3:M3","B — PRÉ-SCREENING","D4A843","1E2761"),
                                    ("N3:W3","C — SCORING /100","2C3E50","FFFFFF"),
                                    ("X3:Y3","D — RÉSULTAT","1A7A3A","FFFFFF"),
                                    ("Z3:AB3","E — NOTES","546E7A","FFFFFF")]:
                ws2.merge_cells(rng); c=ws2[rng.split(":")[0]]
                c.value=lbl; c.font=fb(True,fg,8); c.fill=fx(bg); c.alignment=fc(); c.border=fd()
            ws2.row_dimensions[3].height = 20

            hdrs = ["Produit","Lien fournisseur","Sous-niche","Problème résolu","Cible client","Bénéfice principal",
                    "① Google Trends","Mot-clé #1","Mot-clé #2","② Amazon BSR","③ WinningHunter","④ Minea","Pré-screen",
                    "Demande marché /20","Potentiel pub /15","Marge brute /15","Momentum /10",
                    "Saturation /10","Faisabilité /10","Différenciation /5","Brandabilité /5","Scalabilité /5","PL /5",
                    "Score /100","Verdict","Source","Commentaire","Date"]
            bg_h = {**{i:"1E2761" for i in range(1,7)},**{i:"D4A843" for i in range(7,14)},
                    **{i:"2C3E50" for i in range(14,24)},**{i:"1A7A3A" for i in range(24,26)},
                    **{i:"546E7A" for i in range(26,29)}}
            fg_h = {i:("1E2761" if bg_h.get(i)=="D4A843" else "FFFFFF") for i in range(1,29)}
            for i,h in enumerate(hdrs,1):
                c=ws2[f"{get_column_letter(i)}4"]; c.value=h
                c.font=fb(True,fg_h.get(i,"FFFFFF"),8); c.fill=fx(bg_h.get(i,"2C3E50"))
                c.alignment=fc(); c.border=fd()
            ws2.row_dimensions[4].height = 28

            left_ci = {1,2,4,5,6,8,9,26,27,28}
            for ri, p in enumerate(all_prods, 5):
                row_d = [p["produit"],p["lien_fournisseur"],p["sous_niche"],p["probleme"],p["cible"],p["benefice"],
                         p["gt"],p["gt_kw1"],p["gt_kw2"],p["bsr"],p["wh"],p["minea"],p["pre_screen"],
                         p["demande_marche"],p["potentiel_pub"],p["marge_brute"],p["momentum_tendance"],
                         p["saturation"],p["faisabilite_logistique"],p["differenciation"],p["brandabilite"],
                         p["scalabilite"],p["private_label"],p["score"],p["verdict"],
                         p["source"],p["commentaire"],p["date"]]
                bg_r = "F0F2F8" if ri%2==0 else "FFFFFF"
                for ci,val in enumerate(row_d,1):
                    c=ws2[f"{get_column_letter(ci)}{ri}"]; c.value=val; c.border=fd()
                    c.fill=fx(bg_r); c.font=fb(False,"2C3E50",9)
                    c.alignment=fl() if ci in left_ci else fc()
                vc = VERDICT_COLORS.get(p["verdict"],"2C3E50")
                ws2[f"X{ri}"].font=fb(True,"FFFFFF",11); ws2[f"X{ri}"].fill=fx(vc)
                ws2[f"Y{ri}"].font=fb(True,"FFFFFF",9);  ws2[f"Y{ri}"].fill=fx(vc)
                ws2.row_dimensions[ri].height = 22

            if all_prods:
                ws2.conditional_formatting.add(f"X5:X{4+len(all_prods)}", ColorScaleRule(
                    start_type="num",start_value=0,start_color="C0392B",
                    mid_type="num",mid_value=65,mid_color="F39C12",
                    end_type="num",end_value=100,end_color="27AE60"))

            for i,w in enumerate([22,26,15,20,15,20,12,15,15,12,14,12,18,10,10,10,10,10,10,9,9,9,10,12,22,22,30,14],1):
                ws2.column_dimensions[get_column_letter(i)].width = w
            ws2.freeze_panes = "A5"

            # Summary sheet
            ws3 = wb2.create_sheet("Résumé"); ws3.sheet_view.showGridLines = False
            ws3.merge_cells("A1:E1"); ws3["A1"].value = f"RÉSUMÉ — {vnd}"
            ws3["A1"].font=fb(True,size=12); ws3["A1"].fill=fx("1E2761"); ws3["A1"].alignment=fc()
            ws3.row_dimensions[1].height = 30
            for i,h in enumerate(["Produit","Sous-niche","Score /100","Verdict","Pré-screen"],1):
                c=ws3[f"{get_column_letter(i)}2"]; c.value=h
                c.font=fb(True); c.fill=fx("2C3E50"); c.alignment=fc(); c.border=fd()
            ws3.row_dimensions[2].height = 22
            for ri,p in enumerate(all_prods,3):
                vc=VERDICT_COLORS.get(p["verdict"],"2C3E50")
                for ci,val in enumerate([p["produit"],p["sous_niche"],p["score"],p["verdict"],p["pre_screen"]],1):
                    c=ws3[f"{get_column_letter(ci)}{ri}"]; c.value=val; c.border=fd()
                    c.font=fb(ci in [3,4],"FFFFFF" if ci in [3,4] else "2C3E50",9)
                    c.fill=fx(vc) if ci in [3,4] else fx("F0F2F8" if ri%2==0 else "FFFFFF")
                    c.alignment=fc()
                ws3.row_dimensions[ri].height = 20
            for i,w in enumerate([24,18,12,22,18],1): ws3.column_dimensions[get_column_letter(i)].width = w

            # Save to memory (compatible cloud + local)
            import io
            buffer = io.BytesIO()
            wb2.save(buffer)
            buffer.seek(0)
            slug = (st.session_state.vendeur or "vendeur").replace(" ","_")[:15]
            ts   = datetime.now().strftime("%Y%m%d_%H%M")
            fname = f"ECRA_{slug}_{ts}.xlsx"
            st.download_button(
                f"⬇️ Télécharger — {fname}", buffer,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"✅ {len(all_prods)} produit(s) prêts — clique sur le bouton ci-dessus pour télécharger.")

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1: st.button("← Retour au scoring", on_click=lambda: nav(6))
    with c2:
        if st.button("🗑️ Vider la session complète"):
            st.session_state.produits_liste = []
            reset_product()